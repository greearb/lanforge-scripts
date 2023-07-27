#!/usr/bin/env python3
"""
NAME: lf_db_comparison.py

PURPOSE: lf_db_comparison.py used for comparing the two databases and data manipulations using sql query's

EXAMPLE:
        # Use CLI to compare two different databases:

            ./lf_db_comparison.py --mgr localhost --db1 <path of the database-1> --db2 <path of the database-2> --wct --dp --ap_auto

        # Compare two test runs in a single database using CLI:

            ./lf_db_comparison.py --mgr localhost --database <path of the database> --wct --dp --ap_auto  --index 0,1


SCRIPT_CLASSIFICATION:  Comparison, Excel & Pdf Reporting

SCRIPT_CATEGORIES:   Functional

NOTES:
        * This script can compare the Two different databases which has different test runs in it.

        * This script also have the provision to compare the single database which has multiple test runs in it.

STATUS: BETA RELEASE

VERIFIED_ON:   29-JUN-2023,
             Build Version:  5.4.6
             Kernel Version: 6.2.14+

LICENSE:
          Free to distribute and modify. LANforge systems must be licensed.
          Copyright 2023 Candela Technologies Inc

INCLUDE_IN_README: False
"""

import os
import sys
import time
import importlib
import sqlite3
import logging
import argparse
import openpyxl
import pandas as pd
from datetime import datetime
import datetime
import paramiko
import seaborn as sns
from string import ascii_uppercase
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.append(os.path.join(os.path.abspath(__file__ + "../../../../")))

lf_pdf_report = importlib.import_module("py-scripts.lf_report")
# lf_report = lf_report.lf_report

logger = logging.getLogger(__name__)
lf_logger_config = importlib.import_module("py-scripts.lf_logger_config")


class db_comparison:
    def __init__(self, host, database, data_base1=None, data_base2=None, table_name=None, dp=None, wct=None,
                 ap_auto=None, index=None):
        self.query_df_dict = None
        self.lf_mgr_ip = host
        self.lf_mgr_port = "8080"
        self.lf_mgr_ssh_port = 22
        self.lf_mgr_user = "lanforge"
        self.lf_mgr_pass = "lanforge"
        self.lanforge_gui_version_full = ""
        self.lanforge_gui_git_sha = ""
        self.logger = logging.getLogger(__name__)
        self.conn2 = None
        self.conn1 = None
        self.test_tags = None
        self.directory = None
        self.sub_lists = None
        self.csv_file_name = None
        self.short_description = None
        self.database = database
        self.db1 = data_base1
        self.db2 = data_base2
        self.index = index
        self.dp = dp
        self.wct = wct
        self.ap_auto = ap_auto
        self.table_name = table_name
        if self.database:
            self.db_conn = sqlite3.connect(self.database)
        else:
            self.conn1 = sqlite3.connect(self.db1)
            self.conn2 = sqlite3.connect(self.db2)
        if self.index:
            index_items = self.index[0].split(',')
            if len(index_items) != 2:
                raise argparse.ArgumentTypeError('ERROR : argument --index: expected 2 arguments')
            self.index = [int(item) for item in index_items]

    def building_output_directory(self, directory_name="LRQ_Comparison_Report"):
        now = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S_")  # %Y-%m-%d-%H-h-%m-m-%S-s
        if directory_name:
            self.directory = os.path.join(now + str(directory_name))
            logger.info("Name of the Report Folder: {}".format(self.directory))
        try:
            if not os.path.exists(self.directory):
                os.mkdir(self.directory)
        except Exception as e:
            logger.error("ERROR : The report path is existed but unable to find. Exception raised : {}\n".format(e))

    def checking_data_bases(self, db1, db2):
        # checking if database files are exist
        if not os.path.isfile(db1):
            logger.error(f"Error: File {db1} does not exist\n")
            exit()
        if not os.path.isfile(db2):
            logger.error(f"Error: File {db2} does not exist\n")
            exit()

        # cursor for first database
        cursor1 = self.conn1.cursor()
        # cursor for second database
        cursor2 = self.conn2.cursor()

        # column names from first database
        cursor1.execute(f"SELECT * FROM {self.table_name} LIMIT 1")
        columns1 = [col[0] for col in cursor1.description]
        logger.info(f"First database columns names: {columns1}")

        # column names from second database
        cursor2.execute(f"SELECT * FROM {self.table_name} LIMIT 1")
        columns2 = [col[0] for col in cursor2.description]
        logger.info(f"Second database columns names: {columns2}")

        if columns1 != columns2:
            logger.error("Error: DB1, DB2 column names are not matched.")
            exit()

        cursor1.execute(f"SELECT * FROM {self.table_name}")
        data1 = cursor1.fetchall()

        cursor2.execute(f"SELECT * FROM {self.table_name}")
        data2 = cursor2.fetchall()

        if data1 == data2:
            logger.info("Data is identical (Same) in the given two databases.")
        else:
            logger.info("Data is not identical in the given two databases.")

    def sort_and_merge_db(self, querylist):
        # Query dataframe dictionary
        self.query_df_dict = {
            "single_db_df": [],
            "sorted_single_df": [],
            "db1_df": [],
            "db2_df": [],
            "sorted_db1_df": [],
            "sorted_db2_df": [],
            "merged_df": []
        }
        if querylist:
            if (self.db1 and self.db2) is not None:
                # reading the sql query's from the both databases
                for i in querylist:
                    self.query_df_dict['db1_df'].append(pd.read_sql_query(i, self.conn1))
                    self.query_df_dict['db2_df'].append(pd.read_sql_query(i, self.conn2))
                # sorting the two databases based on test-tags and placing in different list
                for df in range(
                    len(self.query_df_dict['db1_df'])):  #TODO: Need to avoid the length of 'self.query_df_dict['db1_df']'
                    self.query_df_dict["sorted_db1_df"].append(
                        self.query_df_dict["db1_df"][df].replace('\n', '', regex=True).sort_values(by='test-tag', ascending=True))
                    self.query_df_dict["sorted_db2_df"].append(
                        self.query_df_dict["db2_df"][df].replace('\n', '', regex=True).sort_values(by='test-tag', ascending=True))
                for sorted_df in range(len(self.query_df_dict[
                                               'sorted_db1_df'])):  #TODO: Need to avoid the length of 'self.query_df_dict['sorted_db1_df']'
                    self.query_df_dict["merged_df"].append(self.query_df_dict["sorted_db1_df"][sorted_df].merge(
                        self.query_df_dict["sorted_db2_df"][sorted_df], on=['test-tag', 'short-description'],
                        suffixes=('_1', '_2')))
            elif self.database is not None:
                # CONVERTING DF INTO SUB LIST OF DFs BASED ON THE 'Date'
                for i in querylist:
                    self.query_df_dict['single_db_df'].append(pd.read_sql_query(i, self.db_conn))
                new = [[self.query_df_dict['single_db_df'][i].replace('\n', '', regex=True)] for i in range(0, len(self.query_df_dict['single_db_df']))]
                # FETCHING THE INDEX items INTO A LIST TO BREAK THE DATA FRAMES
                sub_lists = []
                for n in range(len(new)):
                    DF = pd.DataFrame(new[n][0])
                    short_description_first_item = DF["short-description"][0]
                    for i, row in new[n][0].iterrows():
                        short_desc = row["short-description"]
                        if short_desc == short_description_first_item:
                            sub_lists.append(i)
                # CONVERTING LIST OF INDEX items into SUB-LISTS
                _index_list = [[]]
                for item in sub_lists:
                    if item == 0:
                        _index_list.append([])
                    else:
                        _index_list[-1].append(item)
                _index_list = [sublist for sublist in _index_list if sublist]

                for sublist in _index_list:
                    if len(sublist) > 1:
                        average_difference = sum(sublist[i] - sublist[i - 1] for i in range(1, len(sublist))) / (
                                len(sublist) - 1)
                        sublist.append(sublist[-1] + int(average_difference))

                # # CONVERTING DF INTO SUB LIST OF DFs BASED ON THE SHORT-DESCRIPTION
                dfs = []
                for i in range(len(new)):
                    last_check = 0
                    for ind in _index_list[i]:
                        dfs.append(new[i][0].loc[last_check:ind - 1])
                        last_check = ind

                # AGAIN CONVERSING LIST WITH IN LIST TO SEPARATE THE DFs
                _index_list = []
                for i in dfs:
                    if i.index[0] == 0:
                        _index_list.append([i])
                    else:
                        _index_list[-1].append(i)
                logger.info("The Total Existing Test Runs in given Data Base : %s" % len(_index_list[0]))
                logger.info("The min existing test runs in given DB : %s" % min([len(i) - 1 for i in _index_list]))

                first_choice = self.index[0]
                second_choice = self.index[1]
                # merging the dataframes and placing in a 'merged_df' list
                for df_length in range(len(_index_list)):
                    self.query_df_dict["merged_df"].append(_index_list[df_length][first_choice].merge(
                        _index_list[df_length][second_choice], on=['test-tag', 'short-description'],
                        suffixes=('_1', '_2')))

                # Define the date format
                date_format = "%d-%b-%Y %H:%M:%S"

                # Convert the Date_1 and Date_2 columns to human-readable date format
                for i in range(len(self.query_df_dict["merged_df"])):
                    self.query_df_dict["merged_df"][i]["Date_1"] = self.query_df_dict["merged_df"][i]["Date_1"].apply(
                        lambda x: datetime.datetime.fromtimestamp(x / 1000).strftime(date_format))
                    self.query_df_dict["merged_df"][i]["Date_2"] = self.query_df_dict["merged_df"][i]["Date_2"].apply(
                        lambda x: datetime.datetime.fromtimestamp(x / 1000).strftime(date_format))
        else:
            logger.info("The List of the query result are empty...")
        return self.query_df_dict

    def percentage_calculation(self, query_dict):
        # Calculating the percentage for the db1-numeric-score & db2-numeric-score and attaching the Comparison (%) values to same dataframe
        percentage_list = []
        query_dict = [df.dropna() for df in query_dict if not df.empty]
        for item in query_dict:
            temp_list = []
            for i in range(len(item['numeric-score_1'])):
                if item['numeric-score_1'][i] != 0:
                    temp_list.append(
                        str(round(abs(((item['numeric-score_2'][i] / item['numeric-score_1'][i]) * 100)), 1)) + "%")
                else:
                    temp_list.append("NaN")
                # if int(item['numeric-score_1'][i]) > int(item['numeric-score_2'][i]):
                #     temp_list.append(
                #         str(round(abs(((item['numeric-score_2'][i] / item['numeric-score_1'][i]) * 100)), 1)) + "%")
                # else:
                #     temp_list.append(
                #         str(round(abs((((item['numeric-score_2'][i] - item['numeric-score_1'][i]) / item['numeric-score_1'][i]) * 100)), 1)) + "%")

            percentage_list.append(temp_list)
            item['Comparison'] = temp_list  # adding the comparison column
            # renaming the data frame keys or column names
            if 'WCT' in item['test-tag'][0]:
                item.rename(columns={'test-tag': 'Test-Tag', 'short-description': 'Short Description - (No of Clients)',
                                     'numeric-score_1': 'Test Run-1 Values', 'numeric-score_2': 'Test Run-2 Values'},
                            inplace=True)
            elif 'DP' in item['test-tag'][0]:
                item.rename(
                    columns={'test-tag': 'Test-Tag', 'short-description': 'Short Description - (NSS-Band-Packet-Size)',
                             'numeric-score_1': 'Test Run-1 Values', 'numeric-score_2': 'Test Run-2 Values'},
                    inplace=True)
            elif 'AP_AUTO' in item['test-tag'][0]:
                item.rename(columns={'test-tag': 'Test-Tag', 'short-description': 'Short Description - (Band Ranges)',
                                     'numeric-score_1': 'Test Run-1 Values', 'numeric-score_2': 'Test Run-2 Values'},
                            inplace=True)
        self.query_df_dict["merged_df"] = query_dict

    def converting_df_to_csv(self, query_df_list):
        if query_df_list:
            for i, df in enumerate(query_df_list):
                if 'WCT' in df['Test-Tag'][0]:
                    with open(f'./{self.directory}/wct.csv', mode='a', newline='') as f:
                        df.to_csv(f, index=True, header=f'Table{i + 1}')
                elif 'DP' in df['Test-Tag'][0]:
                    with open(f'./{self.directory}/dp.csv', mode='a', newline='') as f1:
                        df.to_csv(f1, index=True, header=f'Table{i + 1}')
                elif 'AP_AUTO' in df['Test-Tag'][0]:
                    with open(f'./{self.directory}/ap_auto.csv', mode='a', newline='') as f2:
                        df.to_csv(f2, index=True, header=f'Table{i + 1}')

    def excel_placing(self, query_df_list):
        # creating the different Excel-Sheets for wct, ap_auto, dp & placing the tables side-by-side
        writer_obj = pd.ExcelWriter(f'./{self.directory}/lrq_db_comparison.xlsx', engine='xlsxwriter')

        # Initialize a sublist
        wct_list, dp_list, ap_autolist = [], [], []
        # Iterate over each dataframe in the list L
        for df in query_df_list:
            # Check the condition based on the first value of 'Test-Tag' column
            if 'WCT' in df['Test-Tag'][0]:
                # Append the dataframe to the current sublist
                wct_list.append(df)
            elif 'DP' in df['Test-Tag'][0]:
                dp_list.append(df)
            elif 'AP_AUTO' in df['Test-Tag'][0]:
                ap_autolist.append(df)
        list_dfs = [wct_list, dp_list, ap_autolist]

        # TABLE ARRANGEMENT FOR WIFI CAPACITY WORK SHEET
        row, column = 9, 0
        for i, df in enumerate(wct_list):
            if i > 0:
                row += len(wct_list[i - 1]) + 2
                column = 0
            df.to_excel(writer_obj, sheet_name='LRQ-WiFi_Capacity', index=False, startrow=row, startcol=column)

        # TABLE ARRANGEMENT FOR DATA PLANE WORK SHEET
        row, column = 9, 0
        for i, df in enumerate(dp_list):
            if i > 0:
                row += len(dp_list[i - 1]) + 2
                column = 0
            df.to_excel(writer_obj, sheet_name='LRQ-Data_Plane', index=False, startrow=row, startcol=column)

        # TABLE ARRANGEMENT FOR AP AUTO WORK SHEET
        row, column = 9, 0
        for i, df in enumerate(ap_autolist):
            if i > 0:
                row += len(ap_autolist[i - 1]) + 2
                column = 0
            df.to_excel(writer_obj, sheet_name='LRQ-AP_AUTO', index=False, startrow=row, startcol=column)

        writer_obj.save()

    def excel_adjusting(self, file_name, sheet_name):
        wb = openpyxl.load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Assuming the percentage values start from column B, adjust the range as needed
            for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=11, max_col=11):
                cell = row[0]
                if cell.value is None:
                    pass
                elif cell.value == 'Comparison':
                    pass
                else:
                    percentage = float(cell.value.strip('%'))  # Assuming the percentage values have '%' in them
                    self.set_background_color(cell, percentage)

            for letter in ascii_uppercase:  # ascii_uppercase : A, B, C, D, E, F,......,Z
                max_width = 0
                column_length_list = []
                for row_number in range(1, ws.max_row + 1):
                    length = ws[f'{letter}{row_number}'].value
                    if length is not None:
                        if len(str(length)) > max_width:
                            column_length_list.append(len(str(length)))
                    else:
                        column_length_list.append(10)
                # print("column :", letter)
                # print("width", max(column_length_list))
                ws.column_dimensions[letter].alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[letter].width = max(column_length_list) + 1

            wb.save(file_name)
            wb.close()
        else:
            print(f"None of the sheet {sheet_name} are available for wct, dp, ap_auto...")

    def merge_cells_with_msg(self, file_name, sheet_name, start_column, start_row, end_column, end_row, msg,
                             font_style="DejaVu Serif", font_size=32, font_color="0A9B22"):
        # merge cells
        wb = openpyxl.load_workbook(f'./{self.directory}/lrq_db_comparison.xlsx')
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.merge_cells(f'{start_column}{start_row}:{end_column}{end_row}')

            font_style = Font(name=font_style, size=font_size, bold=True, color=font_color)

            ws.cell(row=int(start_row), column=1).value = msg

            ws.cell(row=int(start_row), column=1).font = font_style

            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

            wb.save(file_name)

            wb.close()
        else:
            print(f"None of the sheet {sheet_name} are available for wct, dp, ap_auto...")

    def set_background_color(self, cell, percentage):
        if percentage >= 90:  # You can define your own conditions here
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                    fill_type="solid")  # Green color for >= 80%
        elif percentage >= 70:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                    fill_type="solid")  # Yellow color for >= 60%
        elif percentage >= 50:
            cell.fill = PatternFill(start_color="E39957", end_color="E39957",
                                    fill_type="solid")  # Orange color for >= 60%
        else:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color for others

    def excel_styling(self, query_df_list):
        # styling the sheets
        self.excel_adjusting(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-WiFi_Capacity')

        self.excel_adjusting(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-Data_Plane')

        self.excel_adjusting(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-AP_AUTO')

        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-WiFi_Capacity',
                                  start_column='A', start_row='2', end_column='K', end_row='3',
                                  msg="WI-FI CAPACITY DATA COMPARISON")
        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-WiFi_Capacity',
                                  start_column='A', start_row='6', end_column='K', end_row='8',
                                  msg="Percentage Calculation Formula : ( Test Run-2 Values / Test Run-1 Values ) * 100 \n"
                                      "Color Indication : GREEN  - % >=  90, YELLOW - % >= 70, ORANGE - % >= 50, RED - % < 50",
                                  font_size=15, font_color="050505")

        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-Data_Plane',
                                  start_column='A', start_row='2', end_column='K', end_row='3',
                                  msg="Data Plane DATA COMPARISON")
        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx',
                                  sheet_name='LRQ-Data_Plane',
                                  start_column='A', start_row='6', end_column='K', end_row='8',
                                  msg="Percentage Calculation Formula : ( Test Run-2 Values / Test Run-1 Values ) * 100 \n"
                                      "Color Indication : GREEN  - % >=  90, YELLOW - % >= 70, ORANGE - % >= 50, RED - % < 50",
                                  font_size=15, font_color="050505")

        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx', sheet_name='LRQ-AP_AUTO',
                                  start_column='A', start_row='2', end_column='K', end_row='3',
                                  msg="AP-AUTO DATA COMPARISON")
        self.merge_cells_with_msg(file_name=f'./{self.directory}/lrq_db_comparison.xlsx',
                                  sheet_name='LRQ-AP_AUTO',
                                  start_column='A', start_row='6', end_column='K', end_row='8',
                                  msg="Percentage Calculation Formula : ( Test Run-2 Values / Test Run-1 Values ) * 100 \n"
                                      "Color Indication : GREEN  - % >=  90, YELLOW - % >= 70, ORANGE - % >= 50, RED - % < 50",
                                  font_size=15, font_color="050505")

        logger.info(f'Excel Report Path: ./{self.directory}/lrq_db_comparison.xlsx')

    def db_querying_with_where_clause(self, column_names, condition='', distinct=False):
        # Querying the databases
        query = None
        db_query = None
        db1_query, db2_query = None, None
        split_column_names = column_names.split(', ')
        adding_quotes = ['"' + item + '"' for item in split_column_names]
        column_names = ', '.join(adding_quotes)
        if not distinct:
            query = ['SELECT ' + column_names + ' FROM ' + self.table_name + ' WHERE ' + condition + ';']
        elif distinct is True:
            query = ['SELECT DISTINCT' + column_names + ' FROM ' + self.table_name + ' WHERE ' + condition + ';']
        # logger.info("Your Data Base Query : {}".format(query))
        if query is not None:
            if (self.conn1 and self.conn2) is not None:
                for i in query:
                    db1_query = pd.read_sql_query(i, self.conn1)
                    db2_query = pd.read_sql_query(i, self.conn2)
                df1 = pd.DataFrame(db1_query)
                df2 = pd.DataFrame(db2_query)
                merged_df = df1.merge(df2, left_index=True, right_index=True, suffixes=('_1', '_2'))
                return merged_df
            elif self.db_conn:
                for i in query:
                    db_query = pd.read_sql_query(i, self.db_conn)
                df = pd.DataFrame(db_query)
                return df
        else:
            logger.info("Query is empty")

    def db_querying_with_limit(self, column_names, condition='', limit=None):
        # Querying the databases
        split_column_names = column_names.split(', ')
        adding_quotes = ['"' + item + '"' for item in split_column_names]
        column_names = ', '.join(adding_quotes)
        query = [
            'SELECT ' + column_names + ' FROM ' + self.table_name + ' WHERE ' + condition + ' LIMIT ' + limit + ';']
        # logger.info(" Your Data Base Query : {}".format(query))
        db1_query, db2_query, db_query = None, None, None
        converted_results = []
        if query is not None:
            if (self.conn1 and self.conn2) is not None:
                for i in query:
                    db1_query = pd.read_sql_query(i, self.conn1)
                    db2_query = pd.read_sql_query(i, self.conn2)
                df1 = pd.DataFrame(db1_query)
                df2 = pd.DataFrame(db2_query)
                if df1.empty and df2.empty:
                    converted_results = []
                else:
                    query_result = [df1.replace(r'\s+', '', regex=True), df2.replace(r'\s+', '', regex=True)]
                    for i, df in enumerate(query_result):
                        result_dict = {}
                        for column in df.columns:
                            result_dict[f"{column}_{i + 1}"] = df[column].values[0]
                        converted_results.append(result_dict)
            elif self.db_conn:
                for i in query:
                    db_query = pd.read_sql_query(i, self.db_conn)
                df = pd.DataFrame(db_query)
                if df.empty:
                    converted_results = []
                else:
                    query_result = [df.replace(r'\s+', '', regex=True)]
                    for i, df in enumerate(query_result):
                        result_dict = {}
                        for column in df.columns:
                            result_dict[f"{column}"] = df[column].values[0]
                        converted_results.append(result_dict)
        else:
            logger.info("Query is empty")
        # logger.info("List of the dictionary values of the two databases : %s" % converted_results)
        return converted_results

    def querying(self):
        # Querying the WCT, AP-AUTO, DP
        list_of_wct_tags, list_of_ap_auto_tags, list_of_dp_tags = [], [], []
        if self.wct:
            # For Wi-Fi Capacity querying
            if (self.db1 and self.db2) is not None:
                wifi_capacity = self.db_querying_with_where_clause(column_names='test-tag',
                                                                   condition='"test-id" == "WiFi Capacity"',
                                                                   distinct=True)
                list_of_wct_tags = sorted(
                    list(set(wifi_capacity['test-tag_1'].unique()) & set(wifi_capacity['test-tag_2'].unique())))
            elif self.database:
                wifi_capacity = self.db_querying_with_where_clause(column_names='test-tag',
                                                                   condition='"test-id" == "WiFi Capacity"',
                                                                   distinct=True)
                list_of_wct_tags = wifi_capacity['test-tag'].tolist()
        if self.ap_auto:
            # For AP Auto querying
            if (self.db1 and self.db2) is not None:
                ap_auto = self.db_querying_with_where_clause(column_names='test-tag',
                                                             condition='"test-id" == "AP Auto"', distinct=True)
                list_of_ap_auto_tags = sorted(
                    list(set(ap_auto['test-tag_1'].unique()) & set(ap_auto['test-tag_2'].unique())))
            elif self.database:
                ap_auto = self.db_querying_with_where_clause(column_names='test-tag',
                                                             condition='"test-id" == "AP Auto"', distinct=True)
                list_of_ap_auto_tags = ap_auto['test-tag'].tolist()
        if self.dp:
            # For Data Plane querying
            if (self.db1 and self.db2) is not None:
                dp_test_tags = self.db_querying_with_where_clause(column_names='test-tag',
                                                                  condition='"test-id" == "Dataplane"', distinct=True)
                list_of_dp_tags = sorted(
                    list(set(dp_test_tags['test-tag_1'].unique()) & set(dp_test_tags['test-tag_2'].unique())))
            elif self.database:
                dp_test_tags = self.db_querying_with_where_clause(column_names='test-tag',
                                                                  condition='"test-id" == "Dataplane"', distinct=True)
                list_of_dp_tags = dp_test_tags['test-tag'].tolist()
        if self.wct or self.ap_auto or self.dp:
            self.test_tags = list_of_wct_tags + list_of_dp_tags + list_of_ap_auto_tags
        logger.info("All Test Tags List : %s" % self.test_tags)

        #  If the list of test-tags for the Wi-Fi Capacity not empty
        if self.test_tags is not None:
            query_results = []
            test_tags = self.test_tags
            break_flag = True
            # Setting up the short_descriptions for WCT
            for i in range(len(test_tags)):
                if 'WCT' in test_tags[i]:
                    if 'UL' in test_tags[i]:
                        self.short_description = 'UL % - % STA'
                    elif 'DL' in test_tags[i]:
                        self.short_description = 'DL % - % STA'
                    else:
                        self.short_description = 'UL+DL % - % STA'
                    break_flag = True
                elif 'AP_AUTO' in test_tags[i]:
                    self.short_description = 'Basic Client Connectivity % %'
                    break_flag = True
                elif 'DP' in test_tags[i]:
                    if (self.db1 and self.db2) is not None:
                        dp_short_desc = self.db_querying_with_where_clause(column_names='test-tag, short-description',
                                                                           condition='"test-id" == "Dataplane"',
                                                                           distinct=True)
                        list_of_short_desc = sorted(list(set(dp_short_desc['short-description_1'].unique()) & set(
                            dp_short_desc['short-description_1'].unique())))
                        slicing_short_description_tag = ['-'.join(str(item).split('-')[0:3]) + '-%' for item in
                                                         list_of_short_desc]
                        sorted_short_description = list(set(slicing_short_description_tag))
                        # self.short_description = 'TCP-%'
                        for short_desc in sorted_short_description:
                            self.short_description = short_desc
                            # querying the db for Wi-fi Capacity
                            query_results.append(
                                'SELECT DISTINCT "test-tag",  "short-description", "kernel", "gui_ver", "gui_build_date", "numeric-score" FROM ' + self.table_name + ' WHERE "test-tag" LIKE \"' +
                                test_tags[i] + '\" and "short-description" LIKE \"' + self.short_description + '\";')
                        break_flag = False
                    elif self.database:
                        dp_short_desc = self.db_querying_with_where_clause(column_names='test-tag, short-description',
                                                                           condition='"test-id" == "Dataplane"',
                                                                           distinct=True)
                        list_of_short_desc = sorted(list(dp_short_desc['short-description'].unique()))
                        slicing_short_description_tag = ['-'.join(str(item).split('-')[0:3]) + '-%' for item in
                                                         list_of_short_desc]
                        sorted_short_description = list(set(slicing_short_description_tag))
                        # self.short_description = 'TCP-%'
                        for short_desc in sorted_short_description:
                            self.short_description = short_desc
                            # querying the db for Wi-fi Capacity
                            query_results.append(
                                'SELECT DISTINCT "test-tag",  "short-description", "Date", "kernel", "gui_ver", "numeric-score" FROM ' + self.table_name + ' WHERE "test-tag" LIKE \"' +
                                test_tags[i] + '\" and "short-description" LIKE \"' + self.short_description + '\" ORDER BY "Date" ASC;')
                        break_flag = False
                # querying the db for Wi-fi Capacity
                if break_flag:
                    if (self.db1 and self.db2) is not None:
                        query_results.append(
                            'SELECT DISTINCT "test-tag",  "short-description", "kernel", "gui_ver", "gui_build_date",  "numeric-score" FROM ' + self.table_name + ' WHERE "test-tag" LIKE \"' +
                            test_tags[i] + '\" and "short-description" LIKE \"' + self.short_description + '\";')
                    elif self.database:
                        query_results.append(
                            'SELECT DISTINCT "test-tag",  "short-description", "Date", "kernel", "gui_ver", "numeric-score" FROM ' + self.table_name + ' WHERE "test-tag" LIKE \"' +
                            test_tags[i] + '\" and "short-description" LIKE \"' + self.short_description + '\" ORDER BY "Date" ASC;')
            # sort and merge the data frames
            self.query_df_dict = self.sort_and_merge_db(querylist=query_results)

            # calculating the percentage
            self.percentage_calculation(query_dict=self.query_df_dict["merged_df"])

            # building the new output directory
            self.building_output_directory()

            # converting the data frames into csv
            self.converting_df_to_csv(query_df_list=self.query_df_dict["merged_df"])

            # placing the tables in Excel sheet side by side
            self.excel_placing(query_df_list=self.query_df_dict["merged_df"])

            # applying the stying for the Excel sheets
            self.excel_styling(query_df_list=self.query_df_dict["merged_df"])

            self.generate_report(dataframes=self.query_df_dict["merged_df"])

    # Fetching the LANForge Full GUI Version Details & lANForge GUI git stash
    def get_lanforge_gui_version(self):
        # creating shh client object we use this object to connect to router
        ssh = paramiko.SSHClient()
        # automatically adds the missing host key
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=self.lf_mgr_ip, port=self.lf_mgr_ssh_port, username=self.lf_mgr_user,
                    password=self.lf_mgr_pass, allow_agent=False, look_for_keys=False, banner_timeout=600)
        stdin, stdout, stderr = ssh.exec_command(
            'curl -H "Accept: application/json" http://{lanforge_ip}:8080 | json_pp  | grep -A 7 "VersionInfo"'.format(
                lanforge_ip=self.lf_mgr_ip))
        self.lanforge_gui_version_full = stdout.readlines()
        self.lanforge_gui_version_full = [line.replace('\n', '') for line in self.lanforge_gui_version_full]
        for element in self.lanforge_gui_version_full:
            if "GitVersion" in element:
                git_sha_str = str(element)
                self.lanforge_gui_git_sha = git_sha_str.split(':', maxsplit=1)[-1].replace(',', '')
                self.logger.info("GitVersion: {}".format(self.lanforge_gui_git_sha))
        ssh.close()
        time.sleep(1)
        return self.lanforge_gui_version_full, self.lanforge_gui_git_sha

    def generate_report(self, dataframes):
        # report = lf_report(_dataframe=dataframe)
        result_directory = f'./{self.directory}/'
        date = str(datetime.datetime.now()).split(",")[0].replace(" ", "-").split(".")[0]
        report_info = {
            "Wifi Capacity": self.wct,
            "Data Plane": self.dp,
            "Ap Auto": self.ap_auto,
        }
        # Fetching the Gui Git Sha & Script Git Sha
        lanforge_gui_git_sha = None

        try:
            lanforge_gui_version_full, lanforge_gui_git_sha = self.get_lanforge_gui_version()
        except Exception:
            logger.error("ERROR: LANForge GUI Version INFO Exception, GUI Ver Please check the lanforge ip")
            # exit(1)

        wct_result = self.db_querying_with_limit(column_names='kernel, gui_ver, dut-model-num',
                                                 condition='"test-id" == "WiFi Capacity"', limit='1')
        dp_result = self.db_querying_with_limit(column_names='kernel, gui_ver, dut-model-num',
                                                condition='"test-id" == "Dataplane"', limit='1')
        ap_auto_result = self.db_querying_with_limit(column_names='kernel, gui_ver, dut-model-num',
                                                     condition='"test-id" == "AP Auto"', limit='1')

        def get_result(result, model_num_1=None, model_num_2=None, model_num=None):
            if result:
                if model_num is None:
                    return result[0][model_num_1], result[1][model_num_2]
                else:
                    return result[0][model_num]
            else:
                if model_num is None:
                    return 'None', 'None'
                else:
                    return 'None'

        gui_info1, gui_info2 = None, None
        if self.db1 and self.db2:
            kernel_1, gui_ver_1, dut_model_num_1 = 'kernel_1', 'gui_ver_1', 'dut-model-num_1'
            kernel_2, gui_ver_2, dut_model_num_2 = 'kernel_2', 'gui_ver_2', 'dut-model-num_2'
            wct_dut1, wct_dut2 = get_result(wct_result, dut_model_num_1, dut_model_num_2)
            dp_dut1, dp_dut2 = get_result(dp_result, dut_model_num_1, dut_model_num_2)
            ap_auto1, ap_auto2 = get_result(ap_auto_result, dut_model_num_1, dut_model_num_2)
            gui_info1 = pd.DataFrame(
                {
                    "Test Run-1 Info": ["WiFi Capacity", "Dataplane", "AP Auto"],
                    # "Kernel Version": [wct_result[0][kernel_1], dp_result[0][kernel_1], ap_auto_result[0][kernel_1]],
                    # "GUI Version": [wct_result[0][gui_ver_1], dp_result[0][gui_ver_1], ap_auto_result[0][gui_ver_1]],
                    "DUT Model": [wct_dut1, dp_dut1, ap_auto1],
                    "GUI git sha": [lanforge_gui_git_sha, lanforge_gui_git_sha, lanforge_gui_git_sha]
                }
            )
            gui_info2 = pd.DataFrame(
                {
                    "Test Run-2 Info": ["WiFi Capacity", "Dataplane", "AP Auto"],
                    # "Kernel Version": [wct_result[1][kernel_2], dp_result[1][kernel_2], ap_auto_result[1][kernel_2]],
                    # "GUI Version": [wct_result[1][gui_ver_2], dp_result[1][gui_ver_2], ap_auto_result[1][gui_ver_2]],
                    "DUT Model": [wct_dut2, dp_dut2, ap_auto2],
                    "GUI git sha": [lanforge_gui_git_sha, lanforge_gui_git_sha, lanforge_gui_git_sha]
                }
            )
        elif self.database:
            kernel, gui_ver, dut_model_num = 'kernel', 'gui_ver', 'dut-model-num'
            wct_dut = get_result(wct_result, model_num=dut_model_num)
            dp_dut = get_result(dp_result, model_num=dut_model_num)
            ap_auto = get_result(ap_auto_result, model_num=dut_model_num)
            gui_info1 = pd.DataFrame(
                {
                    "Test Run-1 Info": ["WiFi Capacity", "Dataplane", "AP Auto"],
                    # "Kernel Version": [wct_result[0][kernel], dp_result[0][kernel], ap_auto_result[0][kernel]],
                    # "GUI Version": [wct_result[0][gui_ver], dp_result[0][gui_ver], ap_auto_result[0][gui_ver]],
                    "DUT Model": [wct_dut, dp_dut, ap_auto],
                    "GUI git sha": [lanforge_gui_git_sha, lanforge_gui_git_sha, lanforge_gui_git_sha]
                }
            )
            gui_info2 = pd.DataFrame(
                {
                    "Test Run-2 Info": ["WiFi Capacity", "Dataplane", "AP Auto"],
                    # "Kernel Version": [wct_result[0][kernel], dp_result[0][kernel], ap_auto_result[0][kernel]],
                    # "GUI Version": [wct_result[0][gui_ver], dp_result[0][gui_ver], ap_auto_result[0][gui_ver]],
                    "DUT Model": [wct_dut, dp_dut, ap_auto],
                    "GUI git sha": [lanforge_gui_git_sha, lanforge_gui_git_sha, lanforge_gui_git_sha]
                }
            )

        report = lf_pdf_report.lf_report(result_directory,
                                         _results_dir_name='lf_db_compare_report',
                                         _output_html="lf_db_compare.html",
                                         _output_pdf="lf_db_compare.pdf")

        # generate output reports
        report.set_title("Percentage Compare Results:")
        report.set_date(date)
        report.build_banner_cover()

        report.set_obj_html("Objective",
                            "The objective of this report is to provide a detailed comparison between two test runs, "
                            "with a primary focus on evaluating the performance of the WiFi capacity, AP auto, and "
                            "data plane functionality. By conducting a comprehensive analysis of these key aspects,"
                            " this report aims to identify any variations or improvements in the test results, "
                            "enabling stakeholders to make informed decisions and take necessary actions to optimize"
                            "the network infrastructure. The report will present a comprehensive assessment of the "
                            "performance metrics, highlighting the strengths and weaknesses of each test run, "
                            "ultimately guiding stakeholders in making effective decisions for enhancing the "
                            "network's overall efficiency and reliability.")
        report.build_objective()

        report.set_table_title("Report Information")
        report.build_table_title()
        report.test_setup_table(value="Database Percentage Results in Report", test_setup_data=report_info)

        report.set_table_title("GUI Overview :")
        report.build_table_title()
        report.set_text("The table below displays the First test run values stored in the database.")
        report.build_text_simple()
        report.set_table_dataframe(gui_info1)
        report.build_table()
        report.set_text("The table below displays the Second test run values stored in the database.")
        report.build_text_simple()
        report.set_table_dataframe(gui_info2)
        report.build_table()

        report.set_table_title("Comparison Tables :")
        report.build_table_title()
        report.set_text(
            "To calculate percentages, use the formula (Test Run-2 Values / Test Run-1 Values) * 100. <div><br>"
            "<u><b>Color Indication</u> :-</b><div><br>"
            " * <font color="'#00FF00'">GREEN</font>  -  % >=  90,  <div><br>"
            " * <font color="'#FFFF00'">YELLOW</font>  -  % >=  70,  <div><br>"
            " * <font color="'#E39957'">ORANGE</font>  -  % >=  50,  <div><br>"
            " * <font color="'#FF0000'">RED</font>     -  % <   50"
        )
        report.build_text_simple()

        # create a dictionary to map keywords to table titles
        title_dict = {'WCT': 'WIFI-CAPACITY', 'DP': 'DATA PLANE', 'AP_AUTO': 'AP AUTO'}

        for i, df in enumerate(dataframes):
            # get the keyword from the Test-Tag column
            keyword = df['Test-Tag'][0].split('_')[0]

            # set the table title based on the keyword
            title = title_dict.get(keyword, 'UNKNOWN')
            # set the table title and dataframe, and build the table
            report.set_table_title(title)
            report.build_table_title()

            def set_pdf_column_color(value):
                x = float(value.strip('%'))
                if x >= 90:
                    color = '#00FF00'  # green
                elif x >= 70:
                    color = '#FFFF00'  # yellow
                elif x >= 50:
                    color = '#E39957'  # orange
                else:
                    color = '#FF0000'  # red
                return f'background:{color}'

            colored_df = df.style.applymap(set_pdf_column_color, subset=['Comparison'])
            report.set_table_dataframe(colored_df)
            report.build_table()

        report_path = report.get_path()
        report_basename = os.path.basename(report_path)
        report_url = './../../' + report_basename
        report.build_link("Current Results Directory", report_url)

        report.build_footer()
        report.write_html()
        report.write_pdf_with_timestamp(_page_size='A4', _orientation='Portrait')


def main():
    parser = argparse.ArgumentParser(prog="lf_db_comparison.py",
                                     formatter_class=argparse.RawTextHelpFormatter,
                                     epilog="""\
                                     DataBase Comparison Using Sql""",
                                     description='''
NAME: lf_db_comparison.py

PURPOSE: lf_db_comparison.py used for comparing the two databases and data manipulations using sql query's

EXAMPLE:
        # Use CLI to compare two different databases:

            ./lf_db_comparison.py --mgr localhost --db1 <path of the database-1> --db2 <path of the database-2> --wct --dp --ap_auto

        # Compare two test runs in a single database using CLI:

            ./lf_db_comparison.py --mgr localhost --database <path of the database> --wct --dp --ap_auto  --index 0,1

SCRIPT_CLASSIFICATION:  Comparison, Excel & Pdf Reporting

SCRIPT_CATEGORIES:   Functional

NOTES:
        * This script can compare the Two different databases which has different test runs in it.

        * This script also have the provision to compare the single database which has multiple test runs in it.

STATUS: BETA RELEASE

VERIFIED_ON:   29-JUN-2023,
             Build Version:  5.4.6
             Kernel Version: 6.2.14+

LICENSE:
          Free to distribute and modify. LANforge systems must be licensed.
          Copyright 2023 Candela Technologies Inc

INCLUDE_IN_README: False

''')
    parser.add_argument('--mgr', help='LANForge ip (USE Only if you want to GUI git sha id in Final report.)',
                        default='localhost')
    parser.add_argument('--database', help='Path to single database file (.db)')
    parser.add_argument('--db1', help='Path to first database file (.db)')
    parser.add_argument('--db2', help='Path to second database file (.db)')
    parser.add_argument('--table_name', help='Name of table to compare', default="qa_table")
    parser.add_argument('--dp', help='To get the DataPlane Comparison', action='store_true')
    parser.add_argument('--wct', help='To get the WiFi Capacity Comparison', action='store_true')
    parser.add_argument('--ap_auto', help='To get the Ap Auto Comparison', action='store_true')
    parser.add_argument('--index',
                        help='Use the \'--index\' argument to specify the index of the test run you want to compare within a single database.\n'
                             'The available options are: \n'
                             '1st TEST RUN : 0, \n'
                             '2nd TEST RUN : 1, \n'
                             '3rd TEST RUN : 2, \n'
                             '.\n'
                             '.\n'
                             'and so on, up to the latest test run. \n'
                             'Example usage:\n'
                             '--index 0, 1 (compares with the first and second test runs.)\n',
                        action='append',
                        default=[])
    # logging configuration:
    parser.add_argument('--log_level', default=None,
                        help='Set logging level: debug | info | warning | error | critical')

    parser.add_argument("--lf_logger_config_json",
                        help="--lf_logger_config_json <json file> , json configuration of logger")

    args = parser.parse_args()

    # set the logger level to debug
    logger_config = lf_logger_config.lf_logger_config()

    if args.log_level:
        logger_config.set_level(level=args.log_level)

    if args.lf_logger_config_json:
        # logger_config.lf_logger_config_json = "lf_logger_config.json"
        logger_config.lf_logger_config_json = args.lf_logger_config_json
        logger_config.load_lf_logger_config()

    logger.debug("Comparing results db1: {db1} db2: {db2} ".format(db1=args.db1, db2=args.db2))

    obj = db_comparison(host=args.mgr,
                        database=args.database,
                        data_base1=args.db1,
                        data_base2=args.db2,
                        table_name=args.table_name,
                        dp=args.dp,
                        wct=args.wct,
                        ap_auto=args.ap_auto,
                        index=args.index)
    # checking the dbs are existed and the data is identical or not
    if not args.database:
        obj.checking_data_bases(db1=args.db1, db2=args.db2)
    # querying
    obj.querying()


if __name__ == "__main__":
    main()
